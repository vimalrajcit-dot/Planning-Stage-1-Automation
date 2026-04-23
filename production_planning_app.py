"""
Production Planning Schedule Generator - Streamlit UI
======================================================
Version: Fully Corrected & Automated with Color Coding + Start Date Normalization + Resource Utilization + TPI Dates Summary
Run with:  streamlit run production_planning_app.py
"""

from struct import pack

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import io
import tempfile
import contextlib
import warnings
warnings.filterwarnings('ignore')

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

import streamlit as st

# ============================================================================

class Config:
    MASTER_FILE_PATH = r"D:\2025-MRO-SP\python\Tag Comprasion code\Production Planning\Production Planning Master.xlsx"
    USER_FILE_PATH = r"D:\2025-MRO-SP\python\Tag Comprasion code\Production Planning\Production Plan.xlsx"
    OUTPUT_FILE_PATH = r"D:\2025-MRO-SP\python\Tag Comprasion code\Production Planning\Production Schedule Output.xlsx"

    WORKING_HOURS_PER_DAY = 8
    WORK_START_HOUR = 8
    WORK_END_HOUR = 16
    WORKING_DAYS = [6,0,1,2,3]  # Sunday-Thursday

    LEAD_TIME_COLUMNS = [
        'Internal Testing','Blasting','Shell Test','Shell Test TPI review','Painting','Paint Curing',
        'Accessory Mounting & Tubing','Calibration & Testing','Document handover & verification by QC (System)',
        'TPI review','Packing'
    ]

    DIRECT_COPY_COLUMNS = [
        'Sales Order No','Customer Name','Tag No.','Work Order No.','Inspection Y/N','Model',
        'Valve size  (inch)','Rating','Body material ','Start Date'
    ]

    SCHEDULE_COLUMN_ORDER = [
        'Sales Order No','Customer Name','Tag No.','Work Order No.','Inspection Y/N','Model',
        'Valve size  (inch)','Rating','Body material ',
        'Internal Testing','Blasting','Shell Test','Shell Test TPI review','Painting','Paint Curing',
        'Accessory Mounting & Tubing','Calibration & Testing','Document handover & verification by QC (System)',
        'TPI review','Packing','Start Date',
        'Internal Testing Completion','Blasting Completion','Shell Test Completion','Shell Test TPI review Completion','TPI witness Start Date','TPI witness Finish Date',
        'Painting Completion','Paint Curing Completion','Accessory Mounting Completion','Calibration Completion',
        'QC Documentation Completion','Final TPI witness Start Date','Final TPI witness Finish Date',
        'Packing Completion','Final Packing TPI witness Start Date','Final Packing TPI witness Finish Date',
        'Dispatch Date','Total Lead Time (hours)'
    ]

    # Stage Load sheet (quantity of stages per day)
    STAGE_LOAD_COLUMNS = ['Date','Internal Testing','Blasting','Shell Test','Shell Test TPI review',
                          'Painting','Paint Curing','Accessory Mounting & Tubing','Calibration & Testing',
                          'Document handover & verification by QC (System)','TPI review','Packing']

    # Stage Load Hours sheet
    STAGE_LOAD_HOURS_COLUMNS = ['Date','Internal Testing','Blasting','Shell Test','Shell Test TPI review',
                                'Painting','Paint Curing','Accessory Mounting & Tubing','Calibration & Testing',
                                'Document handover & verification by QC (System)','TPI review','Packing']

    # Gantt Chart columns
    GANTT_COLUMNS = ['Customer Name','Sales Order No','Tag','Stage','Start','Finish']

    # Resource Utilization columns
    RESOURCE_UTILIZATION_COLUMNS = ['Date','Shift','Stage','Skill Type','Available Hours','Utilized Hours','Utilization %','Extra Hours','Batch Size','Batched Tags']

    # TPI Dates columns
    TPI_DATES_COLUMNS = ['Customer Name', 'Sales Order No', 'Inspection Y/N',
                         'Shell TPI witness Start Date (Min)', 'Shell TPI witness Finish Date (Max)',
                         'Final TPI witness Start Date (Min)', 'Final TPI witness Finish Date (Max)',
                         'Packing TPI witness Start Date (Min)', 'Packing TPI witness Finish Date (Max)']


# ============================================================================

def is_working_day(date):
    if pd.isna(date): return False
    return date.weekday() in Config.WORKING_DAYS

def get_next_working_day(date):
    if pd.isna(date): return None
    next_day = date + timedelta(days=1)
    next_day = next_day.replace(hour=0,minute=0,second=0,microsecond=0)
    while not is_working_day(next_day):
        next_day += timedelta(days=1)
    return next_day

def format_datetime(dt, fmt="%d-%m-%Y %I:%M %p"):
    if pd.isna(dt): return "N/A"
    return dt.strftime(fmt)

# ============================================================================

def calculate_completion_from_hours(start_datetime, lead_time_hours):
    if pd.isna(start_datetime) or pd.isna(lead_time_hours) or lead_time_hours <=0:
        return start_datetime if not pd.isna(start_datetime) else None
    if hasattr(lead_time_hours,'item'): lead_time_hours = lead_time_hours.item()
    lead_time_hours = float(lead_time_hours)

    # Adjust start_datetime within work hours
    if start_datetime.hour < Config.WORK_START_HOUR:
        start_datetime = start_datetime.replace(hour=Config.WORK_START_HOUR, minute=0, second=0, microsecond=0)
    elif start_datetime.hour >= Config.WORK_END_HOUR:
        next_day = get_next_working_day(start_datetime)
        if next_day:
            start_datetime = next_day.replace(hour=Config.WORK_START_HOUR, minute=0, second=0, microsecond=0)

    current_datetime = start_datetime
    remaining_hours = lead_time_hours

    while remaining_hours>0:
        end_of_day = current_datetime.replace(hour=Config.WORK_END_HOUR, minute=0, second=0, microsecond=0)
        hours_left_today = (end_of_day - current_datetime).total_seconds()/3600

        if remaining_hours <= hours_left_today:
            return current_datetime + timedelta(hours=remaining_hours)
        else:
            remaining_hours -= hours_left_today
            next_day = get_next_working_day(current_datetime)
            if next_day:
                current_datetime = next_day.replace(hour=Config.WORK_START_HOUR, minute=0, second=0, microsecond=0)
            else:
                return current_datetime
    return current_datetime

# ============================================================================

def load_master_file(file_path):
    df = pd.read_excel(file_path)
    return df

def load_user_file(file_path):
    df = pd.read_excel(file_path)
    if 'Start Date' in df.columns:
        df['Start Date'] = pd.to_datetime(df['Start Date'], dayfirst=True)
    return df

# ============================================================================

def lookup_lead_times(user_df, master_df):
    user_df['Valve size  (inch)'] = user_df['Valve size  (inch)'].astype(str).str.strip()
    user_df['Rating'] = user_df['Rating'].astype(str).str.strip()
    master_df['Valve size  (inch)'] = master_df['Valve size  (inch)'].astype(str).str.strip()
    master_df['Rating'] = master_df['Rating'].astype(str).str.strip()

    user_keys = user_df['Valve size  (inch)'] + '|' + user_df['Rating']
    master_keys = master_df['Valve size  (inch)'] + '|' + master_df['Rating']
    lookup_dict = {key:idx for idx,key in enumerate(master_keys)}

    for col in Config.LEAD_TIME_COLUMNS:
        if col in master_df.columns: user_df[col]=None

    for idx,user_key in enumerate(user_keys):
        if user_key in lookup_dict:
            master_idx=lookup_dict[user_key]
            for col in Config.LEAD_TIME_COLUMNS:
                if col in master_df.columns:
                    val=master_df.iloc[master_idx][col]
                    if pd.notna(val):
                        val=pd.to_numeric(val, errors='coerce')
                        if pd.notna(val):
                            user_df.at[idx,col]=float(val)
    return user_df

# ============================================================================

def calculate_completion_dates(df):
    activities=[  ('Internal Testing','Internal Testing Completion'),
                ('Blasting','Blasting Completion'), ('Shell Test','Shell Test Completion'),
                ('Shell Test TPI review','Shell Test TPI review Completion'),

 # Painting starts after Shell Test TPI
 ('Painting','Painting Completion'), ('Paint Curing','Paint Curing Completion'),
 ('Accessory Mounting & Tubing','Accessory Mounting Completion'), ('Calibration & Testing','Calibration Completion'),
 ('Document handover & verification by QC (System)','QC Documentation Completion'), ('TPI review','TPI Review Completion'),

    # Packing last
    ('Packing','Packing Completion')

     ]
    for _,comp_col in activities: df[comp_col]=None

    for idx,row in df.iterrows():
        if pd.isna(row.get('Start Date')):
            continue

        # -----------------------------
        # Step 1: Normalize Start Date from Excel
        start_dt = row['Start Date']
        if isinstance(start_dt, str):
            start_dt = pd.to_datetime(start_dt, dayfirst=True, errors='coerce')

        normalized_start = start_dt.replace(second=0, microsecond=0)

        

        # Step 3: Respect working hours
        if normalized_start.hour < Config.WORK_START_HOUR:
            current_start = normalized_start.replace(hour=Config.WORK_START_HOUR, minute=0)
        elif normalized_start.hour >= Config.WORK_END_HOUR:
            next_day = get_next_working_day(normalized_start)
            if next_day:
                current_start = next_day.replace(hour=Config.WORK_START_HOUR, minute=0)
            else:
                current_start = normalized_start
        else:
            current_start = normalized_start

        # -----------------------------
        # Step 4: Calculate all stages
        for lead_col, comp_col in activities:
            lead_time = row.get(lead_col)
            if pd.notna(lead_time) and lead_time > 0:
                completion = calculate_completion_from_hours(current_start, lead_time)
                df.at[idx, comp_col] = completion
                if completion: current_start = completion
            else:
                df.at[idx, comp_col] = current_start

    return df

# ============================================================================

def calculate_milestone_dates(df):

    for idx, row in df.iterrows():

        # --- TPI Witness
        blast = row.get('Blasting Completion')
        shell_tpi = row.get('Shell Test TPI review Completion')

        if pd.notna(blast):
            df.at[idx, 'TPI witness Start Date'] = blast

        if pd.notna(shell_tpi):
            df.at[idx, 'TPI witness Finish Date'] = shell_tpi

        # --- Final TPI (QC + TPI Review duration)
        qc = row.get('QC Documentation Completion')
        tpi_hours = row.get('TPI review')  # this is duration (hours)

        if pd.notna(qc):
            df.at[idx, 'Final TPI witness Start Date'] = qc

            if pd.notna(tpi_hours) and tpi_hours > 0:
                finish = calculate_completion_from_hours(qc, tpi_hours)
                df.at[idx, 'Final TPI witness Finish Date'] = finish
            else:
                df.at[idx, 'Final TPI witness Finish Date'] = qc

        # --- Final Packing TPI (Packing + TPI Review duration)
        pack = row.get('Packing Completion')

        if pd.notna(pack):
            df.at[idx, 'Final Packing TPI witness Start Date'] = pack

            if pd.notna(tpi_hours) and tpi_hours > 0:
                finish = calculate_completion_from_hours(pack, tpi_hours)
                df.at[idx, 'Final Packing TPI witness Finish Date'] = finish
            else:
                df.at[idx, 'Final Packing TPI witness Finish Date'] = pack

        # --- Dispatch (use only Final Packing TPI Finish)
        dispatch_value = df.at[idx, 'Final Packing TPI witness Finish Date']
        df.at[idx, 'Dispatch Date'] = dispatch_value

    return df

# ============================================================================

def generate_tpi_dates_summary(df):
    """
    Generate TPI Dates summary sheet aggregating TPI witness dates by Sales Order No
    """
    tpi_records = []

    # Group by Sales Order No
    grouped = df.groupby('Sales Order No')

    for sales_order, group in grouped:
        # Get Customer Name (first occurrence)
        customer_name = group.iloc[0].get('Customer Name', 'Unknown')

        # Shell TPI witness dates (from Blasting Completion to Shell Test TPI review Completion)
        shell_start_dates = []
        shell_finish_dates = []

        # Final TPI witness dates (from QC Documentation Completion to Final TPI witness Finish Date)
        final_start_dates = []
        final_finish_dates = []

        # Packing TPI witness dates (from Packing Completion to Final Packing TPI witness Finish Date)
        packing_start_dates = []
        packing_finish_dates = []

        for _, row in group.iterrows():
            # Shell TPI
            shell_start = row.get('TPI witness Start Date')
            shell_finish = row.get('TPI witness Finish Date')
            if pd.notna(shell_start):
                shell_start_dates.append(shell_start)
            if pd.notna(shell_finish):
                shell_finish_dates.append(shell_finish)

            # Final TPI
            final_start = row.get('Final TPI witness Start Date')
            final_finish = row.get('Final TPI witness Finish Date')
            if pd.notna(final_start):
                final_start_dates.append(final_start)
            if pd.notna(final_finish):
                final_finish_dates.append(final_finish)

            # Packing TPI
            packing_start = row.get('Final Packing TPI witness Start Date')
            packing_finish = row.get('Final Packing TPI witness Finish Date')
            if pd.notna(packing_start):
                packing_start_dates.append(packing_start)
            if pd.notna(packing_finish):
                packing_finish_dates.append(packing_finish)

        # Calculate min and max for each date range
        shell_start_min = min(shell_start_dates) if shell_start_dates else None
        shell_finish_max = max(shell_finish_dates) if shell_finish_dates else None

        final_start_min = min(final_start_dates) if final_start_dates else None
        final_finish_max = max(final_finish_dates) if final_finish_dates else None

        packing_start_min = min(packing_start_dates) if packing_start_dates else None
        packing_finish_max = max(packing_finish_dates) if packing_finish_dates else None

        tpi_records.append({
            'Customer Name': customer_name,
            'Sales Order No': sales_order,
            'Shell TPI witness Start Date (Min)': shell_start_min,
            'Shell TPI witness Finish Date (Max)': shell_finish_max,
            'Final TPI witness Start Date (Min)': final_start_min,
            'Final TPI witness Finish Date (Max)': final_finish_max,
            'Packing TPI witness Start Date (Min)': packing_start_min,
            'Packing TPI witness Finish Date (Max)': packing_finish_max
        })

    tpi_df = pd.DataFrame(tpi_records)

    # Ensure all columns exist
    for col in Config.TPI_DATES_COLUMNS:
        if col not in tpi_df.columns:
            tpi_df[col] = None

    # Reorder columns
    tpi_df = tpi_df[Config.TPI_DATES_COLUMNS]

    return tpi_df

# ============================================================================

def apply_time_color_coding(workbook, sheet_name="Production Schedule", gantt_sheet_name="Production Gantt Chart"):
    sheets_to_process = [sheet_name, gantt_sheet_name]

    for sheet_name in sheets_to_process:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        headers = [cell.value for cell in ws[1]]

        time_cols = [
            'Start Date','Internal Testing Completion','Blasting Completion','Shell Test Completion','Shell Test TPI review Completion',
            'Painting Completion','Paint Curing Completion','Accessory Mounting Completion','Calibration Completion',
            'QC Documentation Completion','TPI witness Start Date','TPI witness Finish Date','TPI Review Completion',
            'Final TPI witness Start Date','Final TPI witness Finish Date','Final Packing TPI witness Start Date',
            'Final Packing TPI witness Finish Date','Packing Completion','Dispatch Date','Start','Finish', 'Shell Test TPI review Completion'
        ]

        # Deduplicate to avoid processing the same column twice
        time_cols = list(dict.fromkeys(time_cols))

        for col_name in time_cols:
            if col_name not in headers:
                continue
            col_idx = headers.index(col_name) + 1
            for row in range(2, ws.max_row+1):
                cell = ws.cell(row=row, column=col_idx)
                if not isinstance(cell.value, datetime):
                    continue
                hour = cell.value.hour
                if 8 <= hour < 12:
                    cell.fill = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")  # 🔵
                elif 12 <= hour < 16:
                    cell.fill = PatternFill(start_color="F7DC6F", end_color="F7DC6F", fill_type="solid")  # 🟡
                elif 16 <= hour < 20:
                    cell.fill = PatternFill(start_color="EC7063", end_color="EC7063", fill_type="solid")  # 🔴

# ============================================================================

# Manhour Constants
WORK_START = 8
WORK_END = 16
HOURS_PER_DAY = 8

TECH_PEOPLE = 11
SEMI_PEOPLE = 4
UNSKILLED_PEOPLE = 13

TECH_CAPACITY = TECH_PEOPLE * HOURS_PER_DAY
SEMI_CAPACITY = SEMI_PEOPLE * HOURS_PER_DAY
UNSKILLED_CAPACITY = UNSKILLED_PEOPLE * HOURS_PER_DAY

SHELL_TEST_MACHINES = 3
BLAST_MAX_AREA = 6

# Valve Area Table
VALVE_AREA = {
    0.5:0.027,0.75:0.032,1:0.038,1.5:0.045,2:0.081,2.5:0.091,3:0.108,
    4:0.146,6:0.214,8:0.347,10:0.498,12:0.628,14:0.748,16:0.864,
    18:0.960,20:1.076,24:1.554,28:2.014,30:2.489
}

# Stage Resource Type
RESOURCE_TYPE = {
    "Internal Testing":"Tech",
    "Shell Test":"Tech",
    "shell Test TPI review":"Tech",
    "Calibration & Testing":"Tech",
    "Painting":"Semi",
    "Accessory Mounting & Tubing":"Semi",
    "Blasting":"Unskilled",
    "Packing":"Unskilled"
}

# ============================================================================

def paint_optimizer(sizes):
    """
    Optimize painting batches by separating valves by size category
    Returns: (total_batches, batch_details)
    """
    if not sizes:
        return 0, []

    # Separate valves by size category
    small_valves = []      # ≤ 1"
    medium_valves = []     # 1" < size ≤ 4"
    large_valves = []      # > 4"

    for size in sizes:
        if size <= 1:
            small_valves.append(size)
        elif size <= 4:
            medium_valves.append(size)
        else:
            large_valves.append(size)

    batch_details = []
    total_batches = 0

    # Process small valves (capacity 10 per batch)
    if small_valves:
        small_batches = int(np.ceil(len(small_valves) / 10))
        total_batches += small_batches
        batch_details.append({
            'category': 'Small (≤1")',
            'count': len(small_valves),
            'batches': small_batches,
            'capacity': 10,
            'valves': small_valves
        })

    # Process medium valves (capacity 5 per batch)
    if medium_valves:
        medium_batches = int(np.ceil(len(medium_valves) / 5))
        total_batches += medium_batches
        batch_details.append({
            'category': 'Medium (1"-4")',
            'count': len(medium_valves),
            'batches': medium_batches,
            'capacity': 5,
            'valves': medium_valves
        })

    # Process large valves (capacity 2 per batch)
    if large_valves:
        large_batches = int(np.ceil(len(large_valves) / 2))
        total_batches += large_batches
        batch_details.append({
            'category': 'Large (>4")',
            'count': len(large_valves),
            'batches': large_batches,
            'capacity': 2,
            'valves': large_valves
        })

    return total_batches, batch_details

# ============================================================================

def blasting_optimizer(sizes):
    areas = [VALVE_AREA.get(s, 0.1) for s in sizes]

    lots = 1
    area_sum = 0
    lot_tags = []
    batched_info = []
    current_lot_tags = []

    for idx, a in enumerate(areas):
        if area_sum + a <= BLAST_MAX_AREA:
            area_sum += a
            current_lot_tags.append(idx)
        else:
            batched_info.append({'lot': lots, 'tags': current_lot_tags.copy(), 'area': area_sum})
            lots += 1
            area_sum = a
            current_lot_tags = [idx]

    if current_lot_tags:
        batched_info.append({'lot': lots, 'tags': current_lot_tags, 'area': area_sum})

    return lots, batched_info

# ============================================================================

def generate_resource_utilization(df):
    """
    Generate Resource Utilization sheet based on production schedule
    """
    utilization_records = []

    # Get all stages that have resource types defined
    stages = list(RESOURCE_TYPE.keys())

    # For each date in the schedule
    all_dates = set()
    for idx, row in df.iterrows():
        for stage in stages:
            # Map stage names to completion column names
            completion_col_map = {
                'Internal Testing': 'Internal Testing Completion',
                'Blasting': 'Blasting Completion',
                'Shell Test': 'Shell Test Completion',
                'Painting': 'Painting Completion',
                'Accessory Mounting & Tubing': 'Accessory Mounting Completion',
                'Calibration & Testing': 'Calibration Completion',
                'Packing': 'Packing Completion'
            }

            completion_col = completion_col_map.get(stage, f'{stage} Completion')
            start_date = row.get('Start Date')
            completion_date = row.get(completion_col)

            if pd.notna(start_date) and pd.notna(completion_date):
                current_date = start_date.date()
                end_date = completion_date.date()
                while current_date <= end_date:
                    if is_working_day(current_date):
                        all_dates.add(current_date)
                    current_date += timedelta(days=1)

    # Sort dates
    all_dates = sorted(list(all_dates))

    # For each date, collect work orders
    for work_date in all_dates:
        # Collect tags working on each stage on this date
        stage_tags = {stage: [] for stage in stages}
        stage_hours = {stage: [] for stage in stages}
        stage_sizes = {stage: [] for stage in stages}
        stage_time_ranges = {stage: [] for stage in stages}

        for idx, row in df.iterrows():
            tag = row.get('Tag No.', 'Unknown')
            valve_size = row.get('Valve size  (inch)', 0)

            # Convert valve size to float for numeric comparison
            try:
                valve_size_float = float(valve_size) if valve_size != 'N/A' and pd.notna(valve_size) else 0
            except (ValueError, TypeError):
                valve_size_float = 0

            for stage in stages:
                completion_col_map = {
                    'Internal Testing': 'Internal Testing Completion',
                    'Blasting': 'Blasting Completion',
                    'Shell Test': 'Shell Test Completion',
                    'Painting': 'Painting Completion',
                    'Accessory Mounting & Tubing': 'Accessory Mounting Completion',
                    'Calibration & Testing': 'Calibration Completion',
                    'Packing': 'Packing Completion'
                }

                start_col_map = {
                    'Internal Testing': 'Start Date',
                    'Blasting': 'Internal Testing Completion',
                    'Shell Test': 'Blasting Completion',
                    'Painting': 'Shell Test TPI review Completion',
                    'Accessory Mounting & Tubing': 'Painting Completion',
                    'Calibration & Testing': 'Accessory Mounting Completion',
                    'Packing': 'TPI Review Completion'
                }

                start_col = start_col_map.get(stage, 'Start Date')
                completion_col = completion_col_map.get(stage, f'{stage} Completion')

                stage_start = row.get(start_col)
                stage_end = row.get(completion_col)
                stage_hour = row.get(stage, 0)

                # Convert stage_hour to float
                try:
                    stage_hour_float = float(stage_hour) if pd.notna(stage_hour) else 0
                except (ValueError, TypeError):
                    stage_hour_float = 0

                if pd.notna(stage_start) and pd.notna(stage_end):
                    if stage_start.date() <= work_date <= stage_end.date():
                        stage_tags[stage].append(tag)
                        stage_hours[stage].append(stage_hour_float)
                        stage_sizes[stage].append(valve_size_float)
                        stage_time_ranges[stage].append({
                            'start': stage_start,
                            'end': stage_end,
                            'hours': stage_hour_float
                        })

        # Group stages by skill type to calculate shared capacity
        skill_groups = {}
        for stage in stages:
            if stage_tags[stage]:
                skill_type = RESOURCE_TYPE.get(stage, 'Tech')
                if skill_type not in skill_groups:
                    skill_groups[skill_type] = []
                skill_groups[skill_type].append(stage)

        # Calculate utilization for each skill group (shared resources)
        for skill_type, group_stages in skill_groups.items():
            # Get total daily capacity for this skill type
            if skill_type == 'Tech':
                daily_capacity = TECH_CAPACITY
            elif skill_type == 'Semi':
                daily_capacity = SEMI_CAPACITY
            else:  # Unskilled
                daily_capacity = UNSKILLED_CAPACITY

            # Calculate total utilized hours across all stages in this group
            total_utilized_hours = 0
            stage_batch_info = []

            for stage in group_stages:
                # Calculate batch info for this stage
                batch_size_text = ""
                batch_tags_text = ""
                num_valves = len(stage_tags[stage])

                if stage == 'Painting' and stage_sizes[stage]:
                    # Painting batching logic - Separate by size category
                    # Separate valves by size category
                    small_valves_indices = []
                    medium_valves_indices = []
                    large_valves_indices = []

                    for i, size in enumerate(stage_sizes[stage]):
                        if size <= 1:
                            small_valves_indices.append(i)
                        elif size <= 4:
                            medium_valves_indices.append(i)
                        else:
                            large_valves_indices.append(i)

                    # Get tags for each category
                    small_tags = [stage_tags[stage][i] for i in small_valves_indices]
                    medium_tags = [stage_tags[stage][i] for i in medium_valves_indices]
                    large_tags = [stage_tags[stage][i] for i in large_valves_indices]

                    batch_details = []
                    batch_tag_details = []
                    batch_count = 0

                    # Process small valves (capacity 10 per batch)
                    if small_tags:
                        small_batches = int(np.ceil(len(small_tags) / 10))
                        batch_count += small_batches
                        batch_details.append(f"Small (≤1\"): {small_batches} batches (capacity: 10/batch)")

                        # Create batch groups for small valves
                        for i in range(0, len(small_tags), 10):
                            batch_valves = small_tags[i:i+10]
                            batch_tag_details.append(f"Batch {len(batch_tag_details) + 1} - ({', '.join(batch_valves)})")

                    # Process medium valves (capacity 5 per batch)
                    if medium_tags:
                        medium_batches = int(np.ceil(len(medium_tags) / 5))
                        batch_count += medium_batches
                        batch_details.append(f"Medium (1\"-4\"): {medium_batches} batches (capacity: 5/batch)")

                        # Create batch groups for medium valves
                        for i in range(0, len(medium_tags), 5):
                            batch_valves = medium_tags[i:i+5]
                            batch_tag_details.append(f"Batch {len(batch_tag_details) + 1} - ({', '.join(batch_valves)})")

                    # Process large valves (capacity 2 per batch)
                    if large_tags:
                        large_batches = int(np.ceil(len(large_tags) / 2))
                        batch_count += large_batches
                        batch_details.append(f"Large (>4\"): {large_batches} batches (capacity: 2/batch)")

                        # Create batch groups for large valves
                        for i in range(0, len(large_tags), 2):
                            batch_valves = large_tags[i:i+2]
                            batch_tag_details.append(f"Batch {len(batch_tag_details) + 1} - ({', '.join(batch_valves)})")

                    batch_size_text = ' | '.join(batch_details)
                    batch_tags_text = '\n'.join(batch_tag_details)

                elif stage == 'Blasting' and stage_sizes[stage]:
                    # Blasting batching logic
                    lots, batched_info = blasting_optimizer(stage_sizes[stage])
                    batch_size_text = f"Batched into {lots} batches (Area capacity: {BLAST_MAX_AREA}m²)"

                    # Create lot details for Blasting
                    lot_details = []
                    areas = [VALVE_AREA.get(size, 0.1) for size in stage_sizes[stage]]

                    # Create lots based on area optimization
                    current_lot = []
                    current_area = 0
                    lot_num = 1

                    for idx, (tag, area) in enumerate(zip(stage_tags[stage], areas)):
                        if current_area + area <= BLAST_MAX_AREA:
                            current_lot.append(tag)
                            current_area += area
                        else:
                            if current_lot:
                                lot_details.append(f"Lot {lot_num} - ({', '.join(current_lot)}) (Area: {current_area:.2f}m²)")
                                lot_num += 1
                                current_lot = [tag]
                                current_area = area

                    if current_lot:
                        lot_details.append(f"Lot {lot_num} - ({', '.join(current_lot)}) (Area: {current_area:.2f}m²)")

                    batch_tags_text = '\n'.join(lot_details)

                else:
                    # For non-batching stages (Internal Testing, Shell Test, Calibration & Testing, Accessory Mounting, Packing)
                    batch_size_text = f"{num_valves} valves"
                    batch_tags_text = ', '.join(stage_tags[stage])

                stage_batch_info.append({
                    'stage': stage,
                    'batch_size_text': batch_size_text,
                    'batch_tags_text': batch_tags_text,
                    'total_hours': sum(stage_hours[stage])
                })

                # Sum utilized hours
                total_utilized_hours += sum(stage_hours[stage])

            # Cap total utilized hours by daily capacity
            utilized_hours_capped = min(total_utilized_hours, daily_capacity)

            # Calculate overall utilization percentage
            if daily_capacity > 0:
                overall_utilization_pct = (utilized_hours_capped / daily_capacity) * 100
            else:
                overall_utilization_pct = 0

            # Determine which shift(s) the work falls into
            all_start_times = []
            all_end_times = []

            for stage in group_stages:
                for time_range in stage_time_ranges[stage]:
                    if time_range['start'].date() == work_date:
                        all_start_times.append(time_range['start'].hour + time_range['start'].minute/60)
                    if time_range['end'].date() == work_date:
                        all_end_times.append(time_range['end'].hour + time_range['end'].minute/60)

            # Determine shift allocation
            if all_start_times and all_end_times:
                min_start = min(all_start_times)
                max_end = max(all_end_times)

                # Calculate proportional allocation between shifts
                morning_hours = 0
                afternoon_hours = 0

                if min_start < 12:  # Work starts in morning
                    if max_end <= 12:  # All work completed by noon
                        morning_hours = utilized_hours_capped
                    else:  # Work spans both shifts
                        total_duration = max_end - min_start
                        if total_duration > 0:
                            morning_duration = min(12, max_end) - min_start
                            afternoon_duration = max_end - 12
                            morning_hours = (morning_duration / total_duration) * utilized_hours_capped
                            afternoon_hours = (afternoon_duration / total_duration) * utilized_hours_capped
                        else:
                            morning_hours = utilized_hours_capped / 2
                            afternoon_hours = utilized_hours_capped / 2
                else:  # Work starts in afternoon
                    afternoon_hours = utilized_hours_capped

                # Create records for each shift that has work
                shift_hours_map = [
                    ('Morning', morning_hours, daily_capacity / 2),
                    ('Afternoon', afternoon_hours, daily_capacity / 2)
                ]

                for shift_name, shift_utilized, shift_available in shift_hours_map:
                    if shift_utilized > 0:
                        # Calculate extra hours (negative if underutilized, positive if overutilized)
                        extra_hours = shift_utilized - shift_available if shift_utilized > shift_available else 0

                        # Combine batch info for all stages in this skill group
                        combined_batch_size = '\n'.join([f"{info['stage']}: {info['batch_size_text']}" for info in stage_batch_info])
                        combined_batch_tags = '\n\n'.join([f"{info['stage']}:\n{info['batch_tags_text']}" for info in stage_batch_info if info['batch_tags_text']])

                        # Create stage name (combine all stages in this skill group)
                        combined_stage = ' + '.join(group_stages)

                        utilization_records.append({
                            'Date': work_date,
                            'Shift': shift_name,
                            'Stage': combined_stage,
                            'Skill Type': skill_type,
                            'Available Hours': round(shift_available, 1),
                            'Utilized Hours': round(shift_utilized, 1),
                            'Utilization %': round((shift_utilized / shift_available * 100), 1) if shift_available > 0 else 0,
                            'Extra Hours': round(extra_hours, 1),
                            'Batch Size': combined_batch_size,
                            'Batched Tags': combined_batch_tags
                        })
            else:
                # Fallback: single shift record (Morning only)
                shift_name = "Morning"
                shift_available = daily_capacity
                shift_utilized = utilized_hours_capped

                # Calculate extra hours
                extra_hours = shift_utilized - shift_available if shift_utilized > shift_available else 0

                # Combine batch info for all stages in this skill group
                combined_batch_size = '\n'.join([f"{info['stage']}: {info['batch_size_text']}" for info in stage_batch_info])
                combined_batch_tags = '\n\n'.join([f"{info['stage']}:\n{info['batch_tags_text']}" for info in stage_batch_info if info['batch_tags_text']])

                combined_stage = ' + '.join(group_stages)

                utilization_records.append({
                    'Date': work_date,
                    'Shift': shift_name,
                    'Stage': combined_stage,
                    'Skill Type': skill_type,
                    'Available Hours': round(shift_available, 1),
                    'Utilized Hours': round(shift_utilized, 1),
                    'Utilization %': round((shift_utilized / shift_available * 100), 1) if shift_available > 0 else 0,
                    'Extra Hours': round(extra_hours, 1),
                    'Batch Size': combined_batch_size,
                    'Batched Tags': combined_batch_tags
                })

    utilization_df = pd.DataFrame(utilization_records)

    # Update RESOURCE_UTILIZATION_COLUMNS to include Extra Hours
    Config.RESOURCE_UTILIZATION_COLUMNS = ['Date','Shift','Stage','Skill Type','Available Hours','Utilized Hours','Utilization %','Extra Hours','Batch Size','Batched Tags']

    # Reorder columns
    if not utilization_df.empty:
        utilization_df = utilization_df[Config.RESOURCE_UTILIZATION_COLUMNS]

    return utilization_df

# ============================================================================

def apply_resource_utilization_color_coding(worksheet):
    """
    Apply color coding to Resource Utilization sheet based on utilization percentage
    """
    # Find column indices
    headers = []
    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=1, column=col).value
        if cell_value:
            headers.append(cell_value)

    if 'Utilization %' in headers:
        util_col_idx = headers.index('Utilization %') + 1

        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=util_col_idx)
            if cell.value and isinstance(cell.value, (int, float)):
                util_pct = float(cell.value)
                if util_pct > 100:
                    cell.fill = PatternFill(start_color="EC7063", end_color="EC7063", fill_type="solid")  # Red - Over utilization                    
                elif util_pct >= 50:
                    cell.fill = PatternFill(start_color="F7DC6F", end_color="F7DC6F", fill_type="solid")  # Yellow - Good utilization
                elif util_pct < 50:
                    cell.fill = PatternFill(start_color="ABEBC6", end_color="ABEBC6", fill_type="solid")  # Blue - Under utilization
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White - Balanced


# ============================================================================

def apply_stage_load_color_scale(worksheet):
    """
    Apply red-yellow-green color scale to Stage Load data cells.
    Excludes the Date column (col A) and the Total row (last row).
    Green = low load, Yellow = medium, Red = high load (bottleneck).
    """
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    if max_row < 3 or max_col < 2:
        return
    start_cell = f"{get_column_letter(2)}2"
    end_cell = f"{get_column_letter(max_col)}{max_row - 1}"
    rule = ColorScaleRule(
        start_type='min', start_color='63BE7B',                      # Green
        mid_type='percentile', mid_value=50, mid_color='FFEB84',     # Yellow
        end_type='max', end_color='F8696B'                           # Red
    )
    worksheet.conditional_formatting.add(f"{start_cell}:{end_cell}", rule)

# ============================================================================

def apply_sheet_formatting(worksheet):
    """
    Apply professional formatting to a worksheet:
    - Header row: bold, white text, dark-blue fill, centered
    - All data cells: centered with wrap text
    - Auto-fit column widths (capped 10-40 chars)
    """
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Style header row
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Center-align all data cells
    for row in range(2, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row, column=col).alignment = center_align

    # Auto-fit column widths
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            for line in str(cell.value).split('\n'):
                if len(line) > max_length:
                    max_length = len(line)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)

# ============================================================================
def save_output(df, file_path):
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    stages = [
        ('Start Date','Internal Testing Completion','Internal Testing'),
        ('Internal Testing Completion','Blasting Completion','Blasting'),
        ('Blasting Completion','Shell Test Completion','Shell Test'),
        ('Shell Test Completion','Shell Test TPI review Completion','Shell Test TPI review'),
        ('Shell Test TPI review Completion','Painting Completion','Painting'),
        ('Painting Completion','Paint Curing Completion','Paint Curing'),
        ('Paint Curing Completion','Accessory Mounting Completion','Accessory Mounting & Tubing'),
        ('Accessory Mounting Completion','Calibration Completion','Calibration & Testing'),
        ('Calibration Completion','QC Documentation Completion','Document handover & verification by QC (System)'),
        ('QC Documentation Completion','TPI Review Completion','TPI review'),
        ('TPI Review Completion','Packing Completion','Packing')
    ]

    capacity_records = []
    gantt_records = []
    hour_records = []

    for _, row in df.iterrows():
        tag = row.get('Tag No.','Unknown')
        for start_col, end_col, stage in stages:
            start = row.get(start_col)
            end = row.get(end_col)
            if pd.notna(start) and pd.notna(end):
                gantt_records.append({"Customer Name": row.get('Customer Name', 'Unknown'), "Sales Order No": row.get('Sales Order No', 'Unknown'), "Tag": tag, "Stage": stage, "Start": start, "Finish": end})
                # Count each valve ONCE per stage on its start date (true valve-count, not valve-days)
                capacity_records.append({"Date": start.date(), "Stage": stage})
                stage_hours = row.get(stage)
                if pd.notna(stage_hours):
                    hour_records.append({"Stage": stage, "Hours": float(stage_hours), "Date": start.date()})

    gantt_df = pd.DataFrame(gantt_records)
    capacity_df = pd.DataFrame(capacity_records)
    hour_df = pd.DataFrame(hour_records)

    # Stage Load Qty
    stage_load_qty = capacity_df.groupby(['Date','Stage']).size().unstack(fill_value=0).reset_index()

    # ================= Add totals row for Stage Load
    if not stage_load_qty.empty:
        total_row = {'Date': 'Total'}
        for col in stage_load_qty.columns:
            if col != 'Date':
                total_row[col] = stage_load_qty[col].sum()
        stage_load_qty = pd.concat([stage_load_qty, pd.DataFrame([total_row])], ignore_index=True)

    # Stage Load Hours
    stage_load_hours = hour_df.groupby(['Date','Stage'])['Hours'].sum().unstack(fill_value=0).reset_index()

    # ================= Add totals row for Stage Load Hours
    if not stage_load_hours.empty:
        total_row_hours = {'Date': 'Total'}
        for col in stage_load_hours.columns:
            if col != 'Date':
                total_row_hours[col] = stage_load_hours[col].sum()
        stage_load_hours = pd.concat([stage_load_hours, pd.DataFrame([total_row_hours])], ignore_index=True)

    # ================= Generate Resource Utilization
    utilization_df = generate_resource_utilization(df)

    # ================= Generate TPI Dates Summary
    tpi_dates_df = generate_tpi_dates_summary(df)

    # ================= Reorder columns based on Config & fill missing
    # Production Schedule
    export_df = df.copy()
    days_cols = [col for col in export_df.columns if col.endswith("(days)")]
    export_df.drop(columns=days_cols, inplace=True)
    export_df.fillna("N/A", inplace=True)
    export_df = export_df[[col for col in Config.SCHEDULE_COLUMN_ORDER if col in export_df.columns]]

    # Stage Load sheet
    for col in Config.STAGE_LOAD_COLUMNS:
        if col not in stage_load_qty.columns:
            stage_load_qty[col] = 0
    stage_load_qty = stage_load_qty[[col for col in Config.STAGE_LOAD_COLUMNS if col in stage_load_qty.columns]]

    # Stage Load Hours sheet
    for col in Config.STAGE_LOAD_HOURS_COLUMNS:
        if col not in stage_load_hours.columns:
            stage_load_hours[col] = 0
    stage_load_hours = stage_load_hours[[col for col in Config.STAGE_LOAD_HOURS_COLUMNS if col in stage_load_hours.columns]]

    # ================= Write to Excel
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        export_df.to_excel(writer, sheet_name="Production Schedule", index=False)
        stage_load_qty.to_excel(writer, sheet_name="Production Stage Load", index=False)
        stage_load_hours.to_excel(writer, sheet_name="Production Stage Load Hrs", index=False)
        gantt_df.to_excel(writer, sheet_name="Production Gantt Chart", index=False)

        if not utilization_df.empty:
            utilization_df.to_excel(writer, sheet_name="Resource Utilization", index=False)

        if not tpi_dates_df.empty:
            tpi_dates_df.to_excel(writer, sheet_name="TPI Dates", index=False)

        workbook = writer.book

        # Format all datetime columns
        for sheet in workbook.worksheets:
            for col in sheet.columns:
                for cell in col:
                    if isinstance(cell.value, datetime):
                        cell.number_format = "DD-MM-YYYY hh:mm AM/PM"

        # ================= Bold totals row for Stage Load
        stage_load_sheet = writer.sheets["Production Stage Load"]
        totals_row_idx = stage_load_sheet.max_row
        for col in range(1, stage_load_sheet.max_column + 1):
            stage_load_sheet.cell(row=totals_row_idx, column=col).font = Font(bold=True)

        # ================= Bold totals row for Stage Load Hours
        stage_load_hours_sheet = writer.sheets["Production Stage Load Hrs"]
        totals_row_idx_hours = stage_load_hours_sheet.max_row
        for col in range(1, stage_load_hours_sheet.max_column + 1):
            stage_load_hours_sheet.cell(row=totals_row_idx_hours, column=col).font = Font(bold=True)



        # Apply red-yellow-green color scale to Stage Load sheets (excludes Date column & Total row)
        apply_stage_load_color_scale(stage_load_sheet)
        apply_stage_load_color_scale(stage_load_hours_sheet)        

        # Apply color coding on Production Schedule
        apply_time_color_coding(workbook, sheet_name="Production Schedule")

        # Apply color coding on Resource Utilization if it exists
        if "Resource Utilization" in workbook.sheetnames:
            util_sheet = workbook["Resource Utilization"]
            apply_resource_utilization_color_coding(util_sheet)

            # Auto-adjust column widths for Resource Utilization sheet
            for column in util_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                util_sheet.column_dimensions[column_letter].width = adjusted_width

        # Apply color coding on Resource Utilization if it exists
        if "Resource Utilization" in workbook.sheetnames:
            apply_resource_utilization_color_coding(workbook["Resource Utilization"])

            # Unified formatting: colored headers, center alignment, auto-fit for ALL sheets
        for sheet in workbook.worksheets:
            apply_sheet_formatting(sheet)

# ============================================================================

def show_summary(df, utilization_df=None, stage_load_qty=None):
    """
    KPI dashboard summary: totals, averages, bottlenecks, and resource utilization.
    """
    WIDTH = 72

    def section(title):
        print(f"\n┌{(' ' + title + ' ').ljust(WIDTH - 2, '─')}┐")

    def line(label, value):
        content = f"  {label:<36}: {value}"
        if len(content) > WIDTH - 2:
            content = content[:WIDTH - 5] + "..."
        print(f"│{content.ljust(WIDTH - 2)}│")

    def close():
        print("└" + "─" * (WIDTH - 2) + "┘")

    # ═══ Header ═══
    print("\n╔" + "═" * (WIDTH - 2) + "╗")
    print("║" + "PRODUCTION PLANNING DASHBOARD".center(WIDTH - 2) + "║")
    print("╚" + "═" * (WIDTH - 2) + "╝")

    if len(df) == 0:
        print("\n⚠️  No orders to summarize.")
        return

    # ─── Order Overview ───
    section("ORDER OVERVIEW")
    total_tags  = len(df)
    unique_so   = df['Sales Order No'].nunique()  if 'Sales Order No'  in df.columns else 0
    unique_cust = df['Customer Name'].nunique()   if 'Customer Name'   in df.columns else 0

    insp_series = df['Inspection Y/N'].astype(str).str.upper().str.strip() if 'Inspection Y/N' in df.columns else pd.Series([])
    insp_y = int(insp_series.isin(['Y', 'YES']).sum())
    insp_n = int(insp_series.isin(['N', 'NO']).sum())

    line("Total Orders (Tags)",       total_tags)
    line("Unique Sales Orders",       unique_so)
    line("Unique Customers",          unique_cust)
    if total_tags > 0:
        line("Inspection Required (Y)", f"{insp_y} ({insp_y/total_tags*100:.1f}%)")
        line("No Inspection (N)",       f"{insp_n} ({insp_n/total_tags*100:.1f}%)")
    close()

    # ─── Schedule Span ───
    section("SCHEDULE SPAN")
    start_dates    = pd.to_datetime(df['Start Date'],    errors='coerce').dropna() if 'Start Date'    in df.columns else pd.Series([])
    dispatch_dates = pd.to_datetime(df['Dispatch Date'], errors='coerce').dropna() if 'Dispatch Date' in df.columns else pd.Series([])

    if not start_dates.empty:
        line("Earliest Start",    format_datetime(start_dates.min()))
        line("Latest Start",      format_datetime(start_dates.max()))
    if not dispatch_dates.empty:
        line("Earliest Dispatch", format_datetime(dispatch_dates.min()))
        line("Latest Dispatch",   format_datetime(dispatch_dates.max()))
    if not start_dates.empty and not dispatch_dates.empty:
        span_days = (dispatch_dates.max() - start_dates.min()).days
        line("Total Project Duration", f"{span_days} calendar days")
    close()

    # ─── Lead Time Stats ───
    section("LEAD TIME STATS (per order)")
    lead_cols = [c for c in Config.LEAD_TIME_COLUMNS if c in df.columns]
    if lead_cols:
        lead_numeric = df[lead_cols].apply(pd.to_numeric, errors='coerce')
        total_per_order = lead_numeric.sum(axis=1, skipna=True)
        total_per_order = total_per_order[total_per_order > 0]

        if not total_per_order.empty:
            hpd = Config.WORKING_HOURS_PER_DAY
            line("Average Lead Time", f"{total_per_order.mean():.1f} hrs ({total_per_order.mean()/hpd:.1f} days)")
            line("Median Lead Time",  f"{total_per_order.median():.1f} hrs ({total_per_order.median()/hpd:.1f} days)")
            line("Minimum Lead Time", f"{total_per_order.min():.1f} hrs ({total_per_order.min()/hpd:.1f} days)")
            line("Maximum Lead Time", f"{total_per_order.max():.1f} hrs ({total_per_order.max()/hpd:.1f} days)")
    close()

    # ─── Bottlenecks ───
    section("BOTTLENECKS")
    if lead_cols:
        stage_totals = df[lead_cols].apply(pd.to_numeric, errors='coerce').sum()
        if stage_totals.max() > 0:
            line("Busiest Stage (total hours)", f"{stage_totals.idxmax()} — {stage_totals.max():.1f} hrs")
            line("Lightest Stage (total hours)", f"{stage_totals.idxmin()} — {stage_totals.min():.1f} hrs")

    if stage_load_qty is not None and not stage_load_qty.empty:
        qty_only = stage_load_qty[stage_load_qty['Date'].astype(str) != 'Total'].copy()
        if not qty_only.empty:
            stage_cols = [c for c in qty_only.columns if c != 'Date']
            qty_only['_row_total'] = qty_only[stage_cols].sum(axis=1)
            peak_idx   = qty_only['_row_total'].idxmax()
            peak_date  = qty_only.loc[peak_idx, 'Date']
            peak_count = int(qty_only.loc[peak_idx, '_row_total'])
            peak_date_str = peak_date.strftime("%d-%m-%Y") if hasattr(peak_date, 'strftime') else str(peak_date)
            line("Peak Load Day", f"{peak_date_str} ({peak_count} valve-stages)")

            stage_sums = qty_only[stage_cols].sum()
            if stage_sums.max() > 0:
                line("Busiest Stage (valve count)", f"{stage_sums.idxmax()} — {int(stage_sums.max())} valves")
    close()

    # ─── Resource Utilization ───
    if utilization_df is not None and not utilization_df.empty:
        section("RESOURCE UTILIZATION")
        util = utilization_df.copy()
        util['Utilization %'] = pd.to_numeric(util['Utilization %'], errors='coerce')

        for skill in ['Tech', 'Semi', 'Unskilled']:
            sub = util[util['Skill Type'] == skill]
            if not sub.empty:
                line(f"{skill} avg utilization", f"{sub['Utilization %'].mean():.1f}%  (peak {sub['Utilization %'].max():.1f}%)")

        over  = int((util['Utilization %'] > 100).sum())
        under = int((util['Utilization %'] < 50).sum())
        line("Over-utilized shifts (>100%)",  over)
        line("Under-utilized shifts (<50%)",  under)
        close()

    # ─── TPI Witness ───
    section("TPI WITNESS SCHEDULE")
    insp_df = df[insp_series.isin(['Y', 'YES'])] if not insp_series.empty else pd.DataFrame()
    line("Orders requiring TPI", len(insp_df))

    shell_start  = pd.to_datetime(df.get('TPI witness Start Date'),              errors='coerce').dropna() if 'TPI witness Start Date'              in df.columns else pd.Series([])
    final_finish = pd.to_datetime(df.get('Final Packing TPI witness Finish Date'), errors='coerce').dropna() if 'Final Packing TPI witness Finish Date' in df.columns else pd.Series([])

    if not shell_start.empty:
        line("First Shell TPI",         format_datetime(shell_start.min()))
    if not final_finish.empty:
        line("Last Final Packing TPI",  format_datetime(final_finish.max()))
    close()

    print(f"\n📁 Output file: {Config.OUTPUT_FILE_PATH}\n")

# ============================================================================

def main():
    print("=== PRODUCTION PLANNING SCHEDULE GENERATOR ===")
    master_df   = load_master_file(Config.MASTER_FILE_PATH)
    user_df     = load_user_file(Config.USER_FILE_PATH)
    combined_df = lookup_lead_times(user_df.copy(), master_df)
    combined_df = calculate_completion_dates(combined_df)
    combined_df = calculate_milestone_dates(combined_df)

    # Save Excel output
    save_output(combined_df, Config.OUTPUT_FILE_PATH)

    # Compute aggregates independently for the dashboard
    utilization_df = generate_resource_utilization(combined_df)

    stage_load_qty = None
    capacity_records = []
    stages = ['Internal Testing','Blasting','Shell Test','Shell Test TPI review',
              'Painting','Paint Curing','Accessory Mounting & Tubing',
              'Calibration & Testing','Document handover & verification by QC (System)',
              'TPI review','Packing']
    stage_start_map = {
        'Internal Testing': 'Start Date',
        'Blasting': 'Internal Testing Completion',
        'Shell Test': 'Blasting Completion',
        'Shell Test TPI review': 'Shell Test Completion',
        'Painting': 'Shell Test TPI review Completion',
        'Paint Curing': 'Painting Completion',
        'Accessory Mounting & Tubing': 'Paint Curing Completion',
        'Calibration & Testing': 'Accessory Mounting Completion',
        'Document handover & verification by QC (System)': 'Calibration Completion',
        'TPI review': 'QC Documentation Completion',
        'Packing': 'TPI Review Completion',
    }
    for _, row in combined_df.iterrows():
        for stage in stages:
            start = row.get(stage_start_map[stage])
            if pd.notna(start):
                capacity_records.append({'Date': start.date(), 'Stage': stage})
    if capacity_records:
        cap_df = pd.DataFrame(capacity_records)
        stage_load_qty = cap_df.groupby(['Date', 'Stage']).size().unstack(fill_value=0).reset_index()

    # Dashboard
    show_summary(
        combined_df,
        utilization_df=utilization_df,
        stage_load_qty=stage_load_qty,
    )

# ============================================================================
# STREAMLIT UI WRAPPER
# ============================================================================

def run_pipeline_streamlit(master_file_obj, user_file_obj, output_path):
    """
    Streamlit-compatible pipeline that replicates main() but uses uploaded
    file objects for input and a user-chosen output path. All underlying
    logic (lookup_lead_times, calculate_completion_dates,
    calculate_milestone_dates, save_output, generate_resource_utilization,
    show_summary) is called unchanged.
    """
    # Load files from uploaded objects (pd.read_excel accepts file-like objects)
    master_df = load_master_file(master_file_obj)
    user_df   = load_user_file(user_file_obj)

    combined_df = lookup_lead_times(user_df.copy(), master_df)
    combined_df = calculate_completion_dates(combined_df)
    combined_df = calculate_milestone_dates(combined_df)

    # Save Excel output to user-specified temp path
    save_output(combined_df, output_path)

    # Compute aggregates independently for the dashboard
    utilization_df = generate_resource_utilization(combined_df)

    stage_load_qty = None
    capacity_records = []
    stages = ['Internal Testing','Blasting','Shell Test','Shell Test TPI review',
              'Painting','Paint Curing','Accessory Mounting & Tubing',
              'Calibration & Testing','Document handover & verification by QC (System)',
              'TPI review','Packing']
    stage_start_map = {
        'Internal Testing': 'Start Date',
        'Blasting': 'Internal Testing Completion',
        'Shell Test': 'Blasting Completion',
        'Shell Test TPI review': 'Shell Test Completion',
        'Painting': 'Shell Test TPI review Completion',
        'Paint Curing': 'Painting Completion',
        'Accessory Mounting & Tubing': 'Paint Curing Completion',
        'Calibration & Testing': 'Accessory Mounting Completion',
        'Document handover & verification by QC (System)': 'Calibration Completion',
        'TPI review': 'QC Documentation Completion',
        'Packing': 'TPI Review Completion',
    }
    for _, row in combined_df.iterrows():
        for stage in stages:
            start = row.get(stage_start_map[stage])
            if pd.notna(start):
                capacity_records.append({'Date': start.date(), 'Stage': stage})
    if capacity_records:
        cap_df = pd.DataFrame(capacity_records)
        stage_load_qty = cap_df.groupby(['Date', 'Stage']).size().unstack(fill_value=0).reset_index()

    # Capture show_summary console output so we can display it in the UI
    summary_buffer = io.StringIO()
    with contextlib.redirect_stdout(summary_buffer):
        show_summary(
            combined_df,
            utilization_df=utilization_df,
            stage_load_qty=stage_load_qty,
        )
    summary_text = summary_buffer.getvalue()

    return combined_df, utilization_df, stage_load_qty, summary_text


def streamlit_app():
    st.set_page_config(
        page_title="Production Planning Schedule Generator",
        page_icon="🏭",
        layout="wide",
    )

    st.title("🏭 Production Planning Schedule Generator")
    st.markdown(
        "Upload the **Master** lead-time file and the **Production Plan** file, "
        "then click **Generate Schedule** to build the full production schedule, "
        "stage loads, Gantt chart, resource utilization and TPI dates."
    )

    # ─── Sidebar: configuration info ─────────────────────────────────────
    with st.sidebar:
        st.header("⚙️ Configuration")
        st.markdown(f"**Work start:** {Config.WORK_START_HOUR}:00")
        st.markdown(f"**Work end:** {Config.WORK_END_HOUR}:00")
        st.markdown(f"**Hours/day:** {Config.WORKING_HOURS_PER_DAY}")
        st.markdown("**Working days:** Sun–Thu")
        st.markdown("---")
        st.header("👷 Manpower")
        st.markdown(f"**Technicians:** {TECH_PEOPLE}  ({TECH_CAPACITY} hrs/day)")
        st.markdown(f"**Semi-skilled:** {SEMI_PEOPLE}  ({SEMI_CAPACITY} hrs/day)")
        st.markdown(f"**Unskilled:** {UNSKILLED_PEOPLE}  ({UNSKILLED_CAPACITY} hrs/day)")
        st.markdown("---")
        st.header("🛠️ Equipment")
        st.markdown(f"**Shell test machines:** {SHELL_TEST_MACHINES}")
        st.markdown(f"**Blast max area:** {BLAST_MAX_AREA} m²")

    # ─── File uploaders ──────────────────────────────────────────────────
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("📘 Master File")
        master_file = st.file_uploader(
            "Production Planning Master (.xlsx)",
            type=["xlsx"],
            key="master_upload",
        )
    with col2:
        st.subheader("📗 Production Plan")
        user_file = st.file_uploader(
            "Production Plan (.xlsx)",
            type=["xlsx"],
            key="user_upload",
        )

    st.markdown("---")

    # ─── Run button ──────────────────────────────────────────────────────
    run = st.button("🚀 Generate Schedule", type="primary", use_container_width=True)

    if run:
        if master_file is None or user_file is None:
            st.error("❌ Please upload both the Master file and the Production Plan file before running.")
            return

        try:
            with st.spinner("Processing schedule… this may take a moment."):
                # Write output to a temporary file so save_output works unchanged
                tmp_dir = tempfile.mkdtemp()
                output_path = os.path.join(tmp_dir, "Production Schedule Output.xlsx")

                combined_df, utilization_df, stage_load_qty, summary_text = run_pipeline_streamlit(
                    master_file, user_file, output_path
                )

                # Read the saved Excel back for download
                with open(output_path, "rb") as f:
                    excel_bytes = f.read()

            st.success("✅ Schedule generated successfully.")

            # ─── Download button ─────────────────────────────────────────
            st.download_button(
                label="⬇️ Download Excel Output",
                data=excel_bytes,
                file_name="Production Schedule Output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            st.markdown("---")

            # ─── Top-level KPIs ──────────────────────────────────────────
            st.subheader("📊 Key Metrics")
            total_tags  = len(combined_df)
            unique_so   = combined_df['Sales Order No'].nunique()  if 'Sales Order No'  in combined_df.columns else 0
            unique_cust = combined_df['Customer Name'].nunique()   if 'Customer Name'   in combined_df.columns else 0

            insp_series = combined_df['Inspection Y/N'].astype(str).str.upper().str.strip() if 'Inspection Y/N' in combined_df.columns else pd.Series([])
            insp_y = int(insp_series.isin(['Y', 'YES']).sum()) if not insp_series.empty else 0

            dispatch_dates = pd.to_datetime(combined_df['Dispatch Date'], errors='coerce').dropna() if 'Dispatch Date' in combined_df.columns else pd.Series([])
            start_dates    = pd.to_datetime(combined_df['Start Date'],    errors='coerce').dropna() if 'Start Date'    in combined_df.columns else pd.Series([])
            span_days = (dispatch_dates.max() - start_dates.min()).days if (not start_dates.empty and not dispatch_dates.empty) else 0

            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("Total Tags", total_tags)
            m2.metric("Sales Orders", unique_so)
            m3.metric("Customers", unique_cust)
            m4.metric("TPI Required", insp_y)
            m5.metric("Project Span (days)", span_days)

            st.markdown("---")

            # ─── Tabs for each output sheet ─────────────────────────────
            tabs = st.tabs([
                "📋 Production Schedule",
                "📦 Stage Load (Qty)",
                "⏱️ Stage Load (Hrs)",
                "📈 Gantt Data",
                "👥 Resource Utilization",
                "🔍 TPI Dates",
                "🖨️ Console Summary",
            ])

            # Tab 1: Production Schedule
            with tabs[0]:
                display_df = combined_df.copy()
                days_cols = [col for col in display_df.columns if col.endswith("(days)")]
                display_df.drop(columns=days_cols, inplace=True, errors='ignore')
                display_df = display_df[[col for col in Config.SCHEDULE_COLUMN_ORDER if col in display_df.columns]]
                st.dataframe(display_df, use_container_width=True, height=500)

            # Tab 2: Stage Load Qty
            with tabs[1]:
                if stage_load_qty is not None and not stage_load_qty.empty:
                    st.dataframe(stage_load_qty, use_container_width=True, height=500)
                else:
                    st.info("No stage-load data available.")

            # Tab 3: Stage Load Hrs – rebuild quickly from combined_df
            with tabs[2]:
                hour_records = []
                stages_map = [
                    ('Start Date','Internal Testing Completion','Internal Testing'),
                    ('Internal Testing Completion','Blasting Completion','Blasting'),
                    ('Blasting Completion','Shell Test Completion','Shell Test'),
                    ('Shell Test Completion','Shell Test TPI review Completion','Shell Test TPI review'),
                    ('Shell Test TPI review Completion','Painting Completion','Painting'),
                    ('Painting Completion','Paint Curing Completion','Paint Curing'),
                    ('Paint Curing Completion','Accessory Mounting Completion','Accessory Mounting & Tubing'),
                    ('Accessory Mounting Completion','Calibration Completion','Calibration & Testing'),
                    ('Calibration Completion','QC Documentation Completion','Document handover & verification by QC (System)'),
                    ('QC Documentation Completion','TPI Review Completion','TPI review'),
                    ('TPI Review Completion','Packing Completion','Packing'),
                ]
                for _, row in combined_df.iterrows():
                    for start_col, end_col, stage in stages_map:
                        start = row.get(start_col)
                        end   = row.get(end_col)
                        if pd.notna(start) and pd.notna(end):
                            stage_hours = row.get(stage)
                            if pd.notna(stage_hours):
                                hour_records.append({"Stage": stage, "Hours": float(stage_hours), "Date": start.date()})
                if hour_records:
                    hr_df = pd.DataFrame(hour_records)
                    stage_load_hours = hr_df.groupby(['Date','Stage'])['Hours'].sum().unstack(fill_value=0).reset_index()
                    st.dataframe(stage_load_hours, use_container_width=True, height=500)
                else:
                    st.info("No stage-load hours data available.")

            # Tab 4: Gantt
            with tabs[3]:
                gantt_records = []
                for _, row in combined_df.iterrows():
                    tag = row.get('Tag No.','Unknown')
                    for start_col, end_col, stage in stages_map:
                        start = row.get(start_col)
                        end   = row.get(end_col)
                        if pd.notna(start) and pd.notna(end):
                            gantt_records.append({
                                "Customer Name": row.get('Customer Name', 'Unknown'),
                                "Sales Order No": row.get('Sales Order No', 'Unknown'),
                                "Tag": tag, "Stage": stage, "Start": start, "Finish": end
                            })
                if gantt_records:
                    gantt_df = pd.DataFrame(gantt_records)
                    st.dataframe(gantt_df, use_container_width=True, height=500)
                else:
                    st.info("No Gantt data available.")

            # Tab 5: Resource Utilization
            with tabs[4]:
                if utilization_df is not None and not utilization_df.empty:
                    st.dataframe(utilization_df, use_container_width=True, height=500)

                    st.markdown("**Utilization by skill type**")
                    util_num = utilization_df.copy()
                    util_num['Utilization %'] = pd.to_numeric(util_num['Utilization %'], errors='coerce')
                    c1, c2, c3 = st.columns(3)
                    for col, skill in zip([c1, c2, c3], ['Tech', 'Semi', 'Unskilled']):
                        sub = util_num[util_num['Skill Type'] == skill]
                        if not sub.empty:
                            col.metric(
                                f"{skill} avg util.",
                                f"{sub['Utilization %'].mean():.1f}%",
                                f"peak {sub['Utilization %'].max():.1f}%",
                            )
                        else:
                            col.metric(f"{skill} avg util.", "N/A")
                else:
                    st.info("No resource utilization data available.")

            # Tab 6: TPI Dates
            with tabs[5]:
                tpi_df = generate_tpi_dates_summary(combined_df)
                if not tpi_df.empty:
                    st.dataframe(tpi_df, use_container_width=True, height=500)
                else:
                    st.info("No TPI dates data available.")

            # Tab 7: Console summary
            with tabs[6]:
                st.code(summary_text, language="text")

        except Exception as e:
            st.error(f"❌ An error occurred while generating the schedule:\n\n{e}")
            st.exception(e)


# ============================================================================

if __name__ == "__main__":
    # When launched via `streamlit run production_planning_app.py`
    # this block runs the Streamlit UI. To use the original CLI flow,
    # call main() directly instead.
    streamlit_app()
