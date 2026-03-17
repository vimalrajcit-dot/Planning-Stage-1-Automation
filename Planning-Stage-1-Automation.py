"""
Deployment-ready Production Planning Schedule Generator with Streamlit UI
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import warnings
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ====================== CONFIGURATION ==========================
class Config:
    WORKING_HOURS_PER_DAY = 12
    WORK_START_HOUR = 8
    WORK_END_HOUR = 20
    WORKING_DAYS = [6,0,1,2,3]  # Sunday-Thursday

    LEAD_TIME_COLUMNS = [
        'Internal Testing','Blasting','Shell Test','Shell Test TPI review','Painting','Paint Curing',
        'Accessory Mounting & Tubing','Calibration & Testing','Document handover & verification by QC (System)',
        'TPI review','Packing'
    ]

    DIRECT_COPY_COLUMNS = [
        'Sales Order No','Customer Name','Tag No.','Work Order No.','Inspection Y/N',
        'Valve size  (inch)','Rating','Body material ','Start Date'
    ]

    SCHEDULE_COLUMN_ORDER = [
        'Sales Order No','Customer Name','Tag No.','Work Order No.','Inspection Y/N',
        'Valve size  (inch)','Rating','Body material ',
        'Internal Testing','Blasting','Shell Test','Shell Test TPI review','Painting','Paint Curing',
        'Accessory Mounting & Tubing','Calibration & Testing','Document handover & verification by QC (System)',
        'TPI review','Packing','Start Date',
        'Internal Testing Completion','Blasting Completion','Shell Test Completion','Shell Test TPI review Completion','TPI witness Start Date','TPI witness Finish Date',
        'Painting Completion','Paint Curing Completion','Accessory Mounting Completion','Calibration Completion',
        'QC Documentation Completion','Final TPI witness Start Date','Final TPI witness Finish Date',
        'Packing Completion','Final Packing TPI witness Start Date','Final Packing TPI witness Finish Date',
        'Dispatch Date','Total Lead Time (hours)'
    ]

    STAGE_LOAD_COLUMNS = ['Date','Internal Testing','Blasting','Shell Test','Shell Test TPI review',
                          'Painting','Paint Curing','Accessory Mounting & Tubing','Calibration & Testing',
                          'Document handover & verification by QC (System)','TPI review','Packing']

    STAGE_LOAD_HOURS_COLUMNS = ['Date','Internal Testing','Blasting','Shell Test','Shell Test TPI review',
                                'Painting','Paint Curing','Accessory Mounting & Tubing','Calibration & Testing',
                                'Document handover & verification by QC (System)','TPI review','Packing']

    GANTT_COLUMNS = ['Tag','Stage','Start','Finish']

# ====================== HELPER FUNCTIONS ======================
def is_working_day(date):
    if pd.isna(date): return False
    return date.weekday() in Config.WORKING_DAYS

def get_next_working_day(date):
    if pd.isna(date): return None
    next_day = date + timedelta(days=1)
    next_day = next_day.replace(hour=0,minute=0,second=0,microsecond=0)
    while not is_working_day(next_day):
        next_day += timedelta(days=1)
    return next_day

def format_datetime(dt, fmt="%d-%m-%Y %I:%M %p"):
    if pd.isna(dt): return "N/A"
    return dt.strftime(fmt)

def calculate_completion_from_hours(start_datetime, lead_time_hours):
    if pd.isna(start_datetime) or pd.isna(lead_time_hours) or lead_time_hours <=0:
        return start_datetime if not pd.isna(start_datetime) else None
    if hasattr(lead_time_hours,'item'): lead_time_hours = lead_time_hours.item()
    lead_time_hours = float(lead_time_hours)
    if start_datetime.hour < Config.WORK_START_HOUR:
        start_datetime = start_datetime.replace(hour=Config.WORK_START_HOUR, minute=0, second=0, microsecond=0)
    elif start_datetime.hour >= Config.WORK_END_HOUR:
        next_day = get_next_working_day(start_datetime)
        if next_day:
            start_datetime = next_day.replace(hour=Config.WORK_START_HOUR, minute=0, second=0, microsecond=0)
    current_datetime = start_datetime
    remaining_hours = lead_time_hours
    while remaining_hours>0:
        end_of_day = current_datetime.replace(hour=Config.WORK_END_HOUR, minute=0, second=0, microsecond=0)
        hours_left_today = (end_of_day - current_datetime).total_seconds()/3600
        if remaining_hours <= hours_left_today:
            return current_datetime + timedelta(hours=remaining_hours)
        else:
            remaining_hours -= hours_left_today
            next_day = get_next_working_day(current_datetime)
            if next_day:
                current_datetime = next_day.replace(hour=Config.WORK_START_HOUR, minute=0, second=0, microsecond=0)
            else:
                return current_datetime
    return current_datetime

# ====================== DATA PROCESSING ======================
def load_master_file(file_path):
    return pd.read_excel(file_path)

def load_user_file(file_path):
    df = pd.read_excel(file_path)
    if 'Start Date' in df.columns:
        df['Start Date'] = pd.to_datetime(df['Start Date'], dayfirst=True)
    return df

def lookup_lead_times(user_df, master_df):
    user_df['Valve size  (inch)'] = user_df['Valve size  (inch)'].astype(str).str.strip()
    user_df['Rating'] = user_df['Rating'].astype(str).str.strip()
    master_df['Valve size  (inch)'] = master_df['Valve size  (inch)'].astype(str).str.strip()
    master_df['Rating'] = master_df['Rating'].astype(str).str.strip()
    user_keys = user_df['Valve size  (inch)'] + '|' + user_df['Rating']
    master_keys = master_df['Valve size  (inch)'] + '|' + master_df['Rating']
    lookup_dict = {key:idx for idx,key in enumerate(master_keys)}
    for col in Config.LEAD_TIME_COLUMNS:
        if col in master_df.columns: user_df[col]=None
    for idx,user_key in enumerate(user_keys):
        if user_key in lookup_dict:
            master_idx=lookup_dict[user_key]
            for col in Config.LEAD_TIME_COLUMNS:
                if col in master_df.columns:
                    val=master_df.iloc[master_idx][col]
                    if pd.notna(val):
                        val=pd.to_numeric(val, errors='coerce')
                        if pd.notna(val):
                            user_df.at[idx,col]=float(val)
    return user_df

def calculate_completion_dates(df):
    activities=[  ('Internal Testing','Internal Testing Completion'), 
                ('Blasting','Blasting Completion'), ('Shell Test','Shell Test Completion'), 
                ('Shell Test TPI review','Shell Test TPI review Completion'),
                ('Painting','Painting Completion'), ('Paint Curing','Paint Curing Completion'), 
                ('Accessory Mounting & Tubing','Accessory Mounting Completion'), ('Calibration & Testing','Calibration Completion'),
                ('Document handover & verification by QC (System)','QC Documentation Completion'), ('TPI review','TPI Review Completion'),
                ('Packing','Packing Completion')]
    for _,comp_col in activities: df[comp_col]=None
    for idx,row in df.iterrows():
        if pd.isna(row.get('Start Date')): continue
        start_dt = row['Start Date']
        if isinstance(start_dt, str):
            start_dt = pd.to_datetime(start_dt, dayfirst=True, errors='coerce')
        normalized_start = start_dt.replace(second=0, microsecond=0)
        if normalized_start.hour < Config.WORK_START_HOUR:
            current_start = normalized_start.replace(hour=Config.WORK_START_HOUR, minute=0)
        elif normalized_start.hour >= Config.WORK_END_HOUR:
            next_day = get_next_working_day(normalized_start)
            current_start = next_day.replace(hour=Config.WORK_START_HOUR, minute=0) if next_day else normalized_start
        else:
            current_start = normalized_start
        for lead_col, comp_col in activities:
            lead_time = row.get(lead_col)
            if pd.notna(lead_time) and lead_time > 0:
                completion = calculate_completion_from_hours(current_start, lead_time)
                df.at[idx, comp_col] = completion
                if completion: current_start = completion
            else:
                df.at[idx, comp_col] = current_start
    return df

def calculate_milestone_dates(df):
    for idx, row in df.iterrows():
        blast = row.get('Blasting Completion')
        shell_tpi = row.get('Shell Test TPI review Completion')
        if pd.notna(blast): df.at[idx, 'TPI witness Start Date'] = blast
        if pd.notna(shell_tpi): df.at[idx, 'TPI witness Finish Date'] = shell_tpi
        qc = row.get('QC Documentation Completion')
        tpi_hours = row.get('TPI review')
        if pd.notna(qc):
            df.at[idx, 'Final TPI witness Start Date'] = qc
            if pd.notna(tpi_hours) and tpi_hours>0:
                df.at[idx, 'Final TPI witness Finish Date'] = calculate_completion_from_hours(qc, tpi_hours)
            else: df.at[idx, 'Final TPI witness Finish Date'] = qc
        pack = row.get('Packing Completion')
        if pd.notna(pack):
            df.at[idx, 'Final Packing TPI witness Start Date'] = pack
            if pd.notna(tpi_hours) and tpi_hours>0:
                df.at[idx, 'Final Packing TPI witness Finish Date'] = calculate_completion_from_hours(pack, tpi_hours)
            else: df.at[idx, 'Final Packing TPI witness Finish Date'] = pack
        df.at[idx, 'Dispatch Date'] = df.at[idx, 'Final Packing TPI witness Finish Date']
    return df

# ====================== SAVE OUTPUT ==========================
def apply_time_color_coding(workbook, sheet_name="Production Schedule", gantt_sheet_name="Production Gantt Chart"):
    sheets_to_process = [sheet_name, gantt_sheet_name]
    for sheet_name in sheets_to_process:
        if sheet_name not in workbook.sheetnames: continue
        ws = workbook[sheet_name]
        headers = [cell.value for cell in ws[1]]
        time_cols = [
            'Start Date','Internal Testing Completion','Blasting Completion','Shell Test Completion','Shell Test TPI review Completion',
            'Painting Completion','Paint Curing Completion','Accessory Mounting Completion','Calibration Completion',
            'QC Documentation Completion','TPI witness Start Date','TPI witness Finish Date','TPI Review Completion',
            'Final TPI witness Start Date','Final TPI witness Finish Date','Final Packing TPI witness Start Date',
            'Final Packing TPI witness Finish Date','Packing Completion','Dispatch Date','Start','Finish', 'Shell Test TPI review Completion'
        ]
        time_cols = list(dict.fromkeys(time_cols))
        for col_name in time_cols:
            if col_name not in headers: continue
            col_idx = headers.index(col_name)+1
            for row in range(2, ws.max_row+1):
                cell = ws.cell(row=row, column=col_idx)
                if not isinstance(cell.value, datetime): continue
                hour = cell.value.hour
                if 8 <= hour < 12: cell.fill = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
                elif 12 <= hour < 16: cell.fill = PatternFill(start_color="F7DC6F", end_color="F7DC6F", fill_type="solid")
                elif 16 <= hour < 20: cell.fill = PatternFill(start_color="EC7063", end_color="EC7063", fill_type="solid")

def save_output(df, file_path):
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    stages = [
        ('Start Date','Internal Testing Completion','Internal Testing'),
        ('Internal Testing Completion','Blasting Completion','Blasting'),
        ('Blasting Completion','Shell Test Completion','Shell Test'),
        ('Shell Test Completion','Shell Test TPI review Completion','Shell Test TPI review'),
        ('Shell Test TPI review Completion','Painting Completion','Painting'),
        ('Painting Completion','Paint Curing Completion','Paint Curing'),
        ('Paint Curing Completion','Accessory Mounting Completion','Accessory Mounting & Tubing'),
        ('Accessory Mounting Completion','Calibration Completion','Calibration & Testing'),
        ('Calibration Completion','QC Documentation Completion','Document handover & verification by QC (System)'),
        ('QC Documentation Completion','TPI Review Completion','TPI review'),
        ('TPI Review Completion','Packing Completion','Packing')
    ]
    capacity_records = []
    gantt_records = []
    hour_records = []
    for _, row in df.iterrows():
        tag = row.get('Tag No.','Unknown')
        for start_col, end_col, stage in stages:
            start = row.get(start_col)
            end = row.get(end_col)
            if pd.notna(start) and pd.notna(end):
                gantt_records.append({"Tag": tag, "Stage": stage, "Start": start, "Finish": end})
                current = start.date()
                end_date = end.date()
                while current <= end_date:
                    capacity_records.append({"Date": current, "Stage": stage})
                    current += timedelta(days=1)
                stage_hours = row.get(stage)
                if pd.notna(stage_hours):
                    hour_records.append({"Stage": stage, "Hours": float(stage_hours), "Date": start.date()})
    gantt_df = pd.DataFrame(gantt_records)
    capacity_df = pd.DataFrame(capacity_records)
    hour_df = pd.DataFrame(hour_records)
    stage_load_qty = capacity_df.groupby(['Date','Stage']).size().unstack(fill_value=0).reset_index()
    if not stage_load_qty.empty:
        total_row = {'Date': 'Total'}
        for col in stage_load_qty.columns:
            if col != 'Date': total_row[col] = stage_load_qty[col].sum()
        stage_load_qty = pd.concat([stage_load_qty, pd.DataFrame([total_row])], ignore_index=True)
    stage_load_hours = hour_df.groupby(['Date','Stage'])['Hours'].sum().unstack(fill_value=0).reset_index()
    if not stage_load_hours.empty:
        total_row_hours = {'Date': 'Total'}
        for col in stage_load_hours.columns:
            if col != 'Date': total_row_hours[col] = stage_load_hours[col].sum()
        stage_load_hours = pd.concat([stage_load_hours, pd.DataFrame([total_row_hours])], ignore_index=True)
    export_df = df.copy()
    days_cols = [col for col in export_df.columns if col.endswith("(days)")]
    export_df.drop(columns=days_cols, inplace=True)
    export_df.fillna("N/A", inplace=True)
    export_df = export_df[[col for col in Config.SCHEDULE_COLUMN_ORDER if col in export_df.columns]]
    for col in Config.STAGE_LOAD_COLUMNS:
        if col not in stage_load_qty.columns: stage_load_qty[col] = 0
    stage_load_qty = stage_load_qty[[col for col in Config.STAGE_LOAD_COLUMNS if col in stage_load_qty.columns]]
    for col in Config.STAGE_LOAD_HOURS_COLUMNS:
        if col not in stage_load_hours.columns: stage_load_hours[col] = 0
    stage_load_hours = stage_load_hours[[col for col in Config.STAGE_LOAD_HOURS_COLUMNS if col in stage_load_hours.columns]]
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        export_df.to_excel(writer, sheet_name="Production Schedule", index=False)
        stage_load_qty.to_excel(writer, sheet_name="Production Stage Load", index=False)
        stage_load_hours.to_excel(writer, sheet_name="Production Stage Load Hrs", index=False)
        gantt_df.to_excel(writer, sheet_name="Production Gantt Chart", index=False)
        workbook = writer.book
        apply_time_color_coding(workbook, sheet_name="Production Schedule")
        for sheet in workbook.worksheets:
            for col in sheet.columns:
                for cell in col:
                    if isinstance(cell.value, datetime):
                        cell.number_format = "DD-MM-YYYY hh:mm AM/PM"
        stage_load_sheet = writer.sheets["Production Stage Load"]
        totals_row_idx = stage_load_sheet.max_row
        for col in range(1, stage_load_sheet.max_column + 1):
            stage_load_sheet.cell(row=totals_row_idx, column=col).font = Font(bold=True)
        stage_load_hours_sheet = writer.sheets["Production Stage Load Hrs"]
        totals_row_idx_hours = stage_load_hours_sheet.max_row
        for col in range(1, stage_load_hours_sheet.max_column + 1):
            stage_load_hours_sheet.cell(row=totals_row_idx_hours, column=col).font = Font(bold=True)

# ====================== STREAMLIT UI ==========================
st.title("📈 Production Planning Schedule Generator")
st.markdown("Upload Master File and Production Plan Excel files to generate output.")

master_file = st.file_uploader("Upload Production Planning Master File", type=['xlsx'])
user_file = st.file_uploader("Upload Production Plan File", type=['xlsx'])

if master_file and user_file:
    with st.spinner("Processing files... ⏳"):
        master_df = pd.read_excel(master_file)
        user_df = pd.read_excel(user_file)
        if 'Start Date' in user_df.columns:
            user_df['Start Date'] = pd.to_datetime(user_df['Start Date'], dayfirst=True)
        combined_df = lookup_lead_times(user_df.copy(), master_df)
        combined_df = calculate_completion_dates(combined_df)
        combined_df = add_display_columns(combined_df)
        combined_df = calculate_milestone_dates(combined_df)
        output_file_path = "Production_Schedule_Output.xlsx"
        save_output(combined_df, output_file_path)
    st.success("✅ Processing complete!")
    st.subheader("📋 Production Schedule Preview")
    st.dataframe(combined_df.head(20))
    st.subheader("⚡ Production Stage Load Preview")
    stage_load_df = pd.read_excel(output_file_path, sheet_name="Production Stage Load")
    st.dataframe(stage_load_df.head(20))
    st.subheader("⏱️ Production Stage Load Hours Preview")
    stage_load_hours_df = pd.read_excel(output_file_path, sheet_name="Production Stage Load Hrs")
    st.dataframe(stage_load_hours_df.head(20))
    st.subheader("📊 Production